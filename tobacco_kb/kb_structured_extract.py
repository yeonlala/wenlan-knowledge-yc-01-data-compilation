# -*- coding: utf-8 -*-
"""
项目归档资料 → 结构化知识库中间层：Markdown 优先，表格独立 CSV/JSON，原件进 raw/。

抽取链：Docling（可选）→ MarkItDown（可选）→ 按格式回退（docx 段落+表格、pdf 文本、txt/md 直拷贝）。
"""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def _safe_storage_name(name: str, max_len: int = 180) -> str:
    bad = '<>:"/\\|?*\x00-\x1f'
    s = name
    for b in bad:
        s = s.replace(b, "_")
    return s[:max_len] if len(s) > max_len else s


def _stem_idx(idx: int, bundle_key: str) -> str:
    h = hashlib.sha256(bundle_key.encode("utf-8")).hexdigest()[:10]
    return f"{idx:04d}_{h}"


def _heading_level_from_style(style_name: Optional[str]) -> Optional[int]:
    """python-docx 常见样式名 Heading 1..6；无法识别则返回 None（当正文）。"""
    if not style_name or "Heading" not in style_name:
        return None
    tail = style_name.replace("Heading", "").strip()
    if tail.isdigit():
        try:
            return min(6, max(1, int(tail)))
        except ValueError:
            return 2
    return 2


@dataclass
class StructuredExtractResult:
    status: str
    engine: str
    markdown_relative_path: Optional[str] = None
    raw_relative_path: Optional[str] = None  # kb_local 内相对路径，如 raw/xxx.docx
    tables: List[Dict[str, Any]] = field(default_factory=list)
    images: List[Dict[str, Any]] = field(default_factory=list)
    pages: List[Dict[str, Any]] = field(default_factory=list)
    quality: Dict[str, Any] = field(default_factory=dict)
    detail: Optional[str] = None
    legacy_text_relative_path: Optional[str] = None  # 兼容旧字段：部分格式仍写一份纯 txt


def _quality_defaults() -> Dict[str, Any]:
    return {
        "has_text": False,
        "has_tables": False,
        "ocr_used": False,
        "pages_known": False,
    }


def try_docling_markdown(path: Path) -> Optional[Tuple[str, Dict[str, Any]]]:
    """Docling → Markdown。未安装或失败时返回 None。"""
    try:
        from docling.document_converter import DocumentConverter
    except ImportError:
        return None
    try:
        converter = DocumentConverter()
        result = converter.convert(str(path.resolve()))
        doc = result.document
        md = doc.export_to_markdown()
        if not (md or "").strip():
            return None
        meta = {"pages_known": False}
        return (md, meta)
    except Exception:
        return None


def try_markitdown(path: Path) -> Optional[str]:
    try:
        from markitdown import MarkItDown
    except ImportError:
        return None
    try:
        md = MarkItDown()
        res = md.convert(str(path.resolve()))
        txt = res.text_content or ""
        return txt if txt.strip() else None
    except Exception:
        return None


def _iter_docx_blocks(parent: Any):
    from docx.document import Document as DocxDocument
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table, _Cell
    from docx.text.paragraph import Paragraph

    if isinstance(parent, DocxDocument):
        parent_elm = parent.element.body
    elif isinstance(parent, _Cell):
        parent_elm = parent._tc
    else:
        return
    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield ("p", Paragraph(child, parent))
        elif isinstance(child, CT_Tbl):
            yield ("t", Table(child, parent))


def _cell_full_text(cell: Any) -> str:
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table, _Cell
    from docx.text.paragraph import Paragraph

    parts: List[str] = []
    for child in cell._tc.iterchildren():
        if isinstance(child, CT_P):
            parts.append(Paragraph(child, cell).text or "")
        elif isinstance(child, CT_Tbl):
            sub = Table(child, cell)
            rows_txt = []
            for row in sub.rows:
                rows_txt.append(
                    "\t".join(
                        _cell_full_text(c).replace("\n", " ").strip() or (c.text or "").strip()
                        for c in row.cells
                    )
                )
            parts.append("\n".join(rows_txt))
    text = "\n".join(parts).strip()
    return text or (cell.text or "").strip()


def extract_docx_fallback_tables(
    path: Path,
    kb_local: Path,
    stem: str,
    md_path: Path,
) -> Tuple[str, List[Dict[str, Any]]]:
    """
    python-docx：正文按顺序输出 Markdown（段落 / 表格占位），表格另存 CSV+JSON。
    """
    from docx import Document
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    doc = Document(str(path))
    md_lines: List[str] = []
    tables_meta: List[Dict[str, Any]] = []
    table_counter = 0
    tables_dir = kb_local / "extracted_tables"
    tables_dir.mkdir(parents=True, exist_ok=True)

    def walk(parent: Any) -> None:
        nonlocal table_counter
        for kind, block in _iter_docx_blocks(parent):
            if kind == "p" and isinstance(block, Paragraph):
                t = (block.text or "").strip()
                if not t:
                    continue
                st = block.style.name if block.style else None
                hl = _heading_level_from_style(st)
                if hl is not None:
                    md_lines.append(f"{'#' * hl} {t}")
                    md_lines.append("")
                else:
                    md_lines.append(t)
                    md_lines.append("")
            elif kind == "t" and isinstance(block, Table):
                table_counter += 1
                tid = f"{stem}_table_{table_counter:03d}"
                rows_data: List[List[str]] = []
                for row in block.rows:
                    row_cells = []
                    for cell in row.cells:
                        row_cells.append(_cell_full_text(cell))
                    rows_data.append(row_cells)

                csv_name = f"{tid}.csv"
                json_name = f"{tid}.json"
                csv_p = tables_dir / csv_name
                json_p = tables_dir / json_name

                if rows_data:
                    with open(csv_p, "w", encoding="utf-8-sig", newline="") as fp:
                        w = csv.writer(fp)
                        for r in rows_data:
                            w.writerow(r)
                    with open(json_p, "w", encoding="utf-8") as fp:
                        json.dump(
                            {"table_id": tid, "rows": rows_data, "num_rows": len(rows_data)},
                            fp,
                            ensure_ascii=False,
                            indent=2,
                        )

                md_lines.append(f"\n### 表格 {table_counter}\n")
                md_lines.append(f"<!-- structured_table: {tid} -->")
                md_lines.append(f"- CSV: `extracted_tables/{csv_name}`")
                md_lines.append(f"- JSON: `extracted_tables/{json_name}`")
                if rows_data:
                    md_lines.append("")
                    md_lines.append(_rows_to_pipe_markdown(rows_data))
                    md_lines.append("")

                tables_meta.append(
                    {
                        "table_id": tid,
                        "csv_path": f"extracted_tables/{csv_name}",
                        "json_path": f"extracted_tables/{json_name}",
                        "page": None,
                        "caption": None,
                        "num_rows": len(rows_data),
                    }
                )

    walk(doc)
    body = "\n".join(md_lines).strip() + "\n"
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text(body, encoding="utf-8")
    return body, tables_meta


def _rows_to_pipe_markdown(rows: List[List[str]]) -> str:
    if not rows:
        return ""
    esc = lambda s: (s or "").replace("|", "\\|").replace("\n", " ")
    header = rows[0]
    lines = ["| " + " | ".join(esc(c) for c in header) + " |", "| " + " | ".join("---" for _ in header) + " |"]
    for r in rows[1:]:
        lines.append("| " + " | ".join(esc(c) for c in r) + " |")
    return "\n".join(lines)


def extract_xlsx_structured(
    path: Path,
    kb_local: Path,
    stem: str,
    md_path: Path,
) -> Tuple[str, List[Dict[str, Any]]]:
    """每个 sheet → CSV/JSON + Markdown 内嵌 pipe 表。"""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    md_lines: List[str] = []
    tables_meta: List[Dict[str, Any]] = []
    tables_dir = kb_local / "extracted_tables"
    tables_dir.mkdir(parents=True, exist_ok=True)

    for si, sheetname in enumerate(wb.sheetnames):
        ws = wb[sheetname]
        rows_data: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append(["" if c is None else str(c) for c in row])
        while rows_data and all(not (x or "").strip() for x in rows_data[-1]):
            rows_data.pop()

        safe_sheet = (sheetname or f"sheet_{si + 1}").replace("\n", " ")
        if not rows_data:
            md_lines.append(f"\n## {safe_sheet}\n\n_（无数据行）_\n")
            continue

        tid = f"{stem}_sheet_{si + 1:03d}"
        csv_name = f"{tid}.csv"
        json_name = f"{tid}.json"
        csv_p = tables_dir / csv_name
        json_p = tables_dir / json_name

        if rows_data:
            with open(csv_p, "w", encoding="utf-8-sig", newline="") as fp:
                w = csv.writer(fp)
                for r in rows_data:
                    w.writerow(r)
            with open(json_p, "w", encoding="utf-8") as fp:
                json.dump(
                    {"table_id": tid, "sheet": sheetname, "rows": rows_data, "num_rows": len(rows_data)},
                    fp,
                    ensure_ascii=False,
                    indent=2,
                )

        md_lines.append(f"\n## {safe_sheet}\n")
        md_lines.append(f"<!-- structured_table: {tid} -->")
        md_lines.append(f"- CSV: `extracted_tables/{csv_name}`")
        md_lines.append(f"- JSON: `extracted_tables/{json_name}`")
        if rows_data:
            md_lines.append("")
            md_lines.append(_rows_to_pipe_markdown(rows_data))
            md_lines.append("")

        tables_meta.append(
            {
                "table_id": tid,
                "csv_path": f"extracted_tables/{csv_name}",
                "json_path": f"extracted_tables/{json_name}",
                "page": None,
                "caption": safe_sheet,
                "num_rows": len(rows_data),
            }
        )

    wb.close()
    body = "\n".join(md_lines).strip() + "\n"
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text(body, encoding="utf-8")
    return body, tables_meta


def extract_pdf_fallback(path: Path, md_path: Path) -> Tuple[str, Dict[str, Any]]:
    """pypdf 按页 → Markdown；正文仍为空时用 pdfplumber 再试。"""
    chunks: List[str] = []

    try:
        from pypdf import PdfReader

        r = PdfReader(str(path))
        for i, page in enumerate(r.pages):
            t = (page.extract_text() or "").strip()
            if t:
                chunks.append(f"## 第 {i + 1} 页\n\n{t}")
    except ImportError:
        pass

    if not chunks:
        try:
            import pdfplumber

            with pdfplumber.open(str(path)) as pdf:
                for i, page in enumerate(pdf.pages):
                    t = (page.extract_text() or "").strip()
                    if t:
                        chunks.append(f"## 第 {i + 1} 页\n\n{t}")
        except ImportError:
            pass

    body = "\n\n".join(chunks)
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text(body + "\n", encoding="utf-8")
    q = _quality_defaults()
    q["has_text"] = bool(body.strip())
    q["pages_known"] = True
    return body, q


def extract_document_to_kb_bundle(
    source_path: Path,
    kb_local: Path,
    file_index: int,
    bundle_key: str,
    *,
    raw_source_path: Optional[Path] = None,
) -> StructuredExtractResult:
    """
    统一入口：写入 kb_local/raw、extracted_markdown（及必要时 extracted_tables），返回结构化结果。

    raw_source_path：若抽取用的是转换后的临时文件（如 .doc→.docx），此处传入 **原始** 路径以便复制到 raw/。
    """
    ext = source_path.suffix.lower()
    stem = _stem_idx(file_index, bundle_key)
    raw_dir = kb_local / "raw"
    md_dir = kb_local / "extracted_markdown"
    img_dir = kb_local / "extracted_images"
    raw_dir.mkdir(parents=True, exist_ok=True)
    md_dir.mkdir(parents=True, exist_ok=True)
    img_dir.mkdir(parents=True, exist_ok=True)

    copy_from = raw_source_path if raw_source_path is not None else source_path
    safe_bn = _safe_storage_name(bundle_key)
    raw_dest = raw_dir / safe_bn
    try:
        shutil.copy2(copy_from, raw_dest)
    except OSError:
        raw_dest = raw_dir / f"{stem}_{Path(bundle_key).name}"
        shutil.copy2(copy_from, raw_dest)

    raw_rel = f"raw/{raw_dest.name}"

    md_out = md_dir / f"{stem}.md"
    q = _quality_defaults()

    # ---------- DOCX ----------
    if ext == ".docx":
        dl = try_docling_markdown(source_path)
        if dl is not None:
            md_text, dm = dl
            md_out.write_text(md_text + "\n", encoding="utf-8")
            q["has_text"] = bool(md_text and md_text.strip())
            # Docling 已在 md 内含表格；另尝试拆表文件：若无 pandas 则跳过
            tables_out: List[Dict[str, Any]] = []
            return StructuredExtractResult(
                status="ok",
                engine="docling",
                raw_relative_path=raw_rel,
                markdown_relative_path=f"extracted_markdown/{md_out.name}",
                tables=tables_out,
                quality={**q, **dm},
            )

        mk = try_markitdown(source_path)
        if mk is not None and mk.strip():
            md_out.write_text(mk + "\n", encoding="utf-8")
            q["has_text"] = True
            return StructuredExtractResult(
                status="ok",
                engine="markitdown",
                raw_relative_path=raw_rel,
                markdown_relative_path=f"extracted_markdown/{md_out.name}",
                tables=[],
                quality=q,
            )

        _, tables_meta = extract_docx_fallback_tables(source_path, kb_local, stem, md_out)
        q["has_tables"] = len(tables_meta) > 0
        q["has_text"] = md_out.is_file() and md_out.stat().st_size > 0
        return StructuredExtractResult(
            status="ok",
            engine="docx_structured_fallback",
            raw_relative_path=raw_rel,
            markdown_relative_path=f"extracted_markdown/{md_out.name}",
            tables=tables_meta,
            quality=q,
        )

    # ---------- PDF ----------
    if ext == ".pdf":
        dl = try_docling_markdown(source_path)
        if dl is not None:
            md_text, dm = dl
            md_out.write_text(md_text + "\n", encoding="utf-8")
            q["has_text"] = bool(md_text and md_text.strip())
            return StructuredExtractResult(
                status="ok",
                engine="docling",
                raw_relative_path=raw_rel,
                markdown_relative_path=f"extracted_markdown/{md_out.name}",
                quality={**q, **dm},
            )

        mk = try_markitdown(source_path)
        if mk is not None and mk.strip():
            md_out.write_text(mk + "\n", encoding="utf-8")
            q["has_text"] = True
            return StructuredExtractResult(
                status="ok",
                engine="markitdown",
                raw_relative_path=raw_rel,
                markdown_relative_path=f"extracted_markdown/{md_out.name}",
                quality=q,
            )

        _, pq = extract_pdf_fallback(source_path, md_out)
        return StructuredExtractResult(
            status="ok",
            engine="pdf_text_fallback",
            raw_relative_path=raw_rel,
            markdown_relative_path=f"extracted_markdown/{md_out.name}",
            quality=pq,
        )

    # ---------- 纯文本 / 已有 md ----------
    if ext in {".txt", ".md"}:
        raw = source_path.read_text(encoding="utf-8", errors="replace")
        md_out.write_text(raw + ("\n" if not raw.endswith("\n") else ""), encoding="utf-8")
        q["has_text"] = bool(raw.strip())
        return StructuredExtractResult(
            status="ok",
            engine="text_plain",
            raw_relative_path=raw_rel,
            markdown_relative_path=f"extracted_markdown/{md_out.name}",
            quality=q,
        )

    # ---------- Excel ----------
    if ext == ".xlsx":
        try:
            _, tables_meta = extract_xlsx_structured(source_path, kb_local, stem, md_out)
            q["has_tables"] = len(tables_meta) > 0
            q["has_text"] = md_out.is_file() and md_out.stat().st_size > 0
            return StructuredExtractResult(
                status="ok",
                engine="openpyxl_structured",
                raw_relative_path=raw_rel,
                markdown_relative_path=f"extracted_markdown/{md_out.name}",
                tables=tables_meta,
                quality=q,
            )
        except Exception as e:
            return StructuredExtractResult(
                status="failed",
                engine="openpyxl_structured",
                raw_relative_path=raw_rel,
                detail=str(e),
                quality=q,
            )

    # ---------- .doc：仅复制 raw；若能 LO 转 txt 再包一层 md（沿用旧逻辑可在外部调）----------
    if ext == ".doc":
        return StructuredExtractResult(
            status="skipped_format",
            engine="none",
            raw_relative_path=raw_rel,
            detail="请先 LibreOffice/Word 转为 docx 再结构化；原件已在 raw/",
            quality=q,
        )

    return StructuredExtractResult(
        status="skipped_format",
        engine="none",
        raw_relative_path=raw_rel,
        detail=f"unsupported extension {ext}",
        quality=q,
    )


def structured_result_to_manifest_extraction(res: StructuredExtractResult) -> Dict[str, Any]:
    """转为 manifest 中 extraction 对象（schema 2.x）。"""
    out: Dict[str, Any] = {
        "status": res.status,
        "engine": res.engine,
        "backend": res.engine,
        "source_file_ref": res.raw_relative_path,
        "markdown_path": res.markdown_relative_path,
        "tables": res.tables,
        "images": res.images,
        "pages": res.pages,
        "quality": res.quality,
        "text_relative_path": res.legacy_text_relative_path,
        "detail": res.detail,
    }
    if res.status != "ok":
        out["markdown_path"] = None
    return out
