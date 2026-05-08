import html
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

from tobacco_kb.naming_convention import (
    DATE_SEGMENT_RE,
    VALID_CATEGORIES,
    VALID_EXTENSIONS,
    VALID_SECRET_LEVELS,
    VERSION_SEGMENT_RE,
    try_parse_confirmation_bucket_stem,
    try_parse_simplified_stem,
)
from tobacco_kb.acceptance_config import (
    ACCEPTANCE_THRESHOLDS,
    PHASE1_RULES,
    TOP_BUCKET_RULES,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少 openpyxl，请先安装：pip install openpyxl")
    sys.exit(1)

# =========================
# 1. 基础配置
# =========================

# 与「01_管理确认/确认单生成器」等工具目录：整目录不参与机检（不扫描其下任何文件）
CONFIRM_GENERATOR_DIRNAME = "确认单生成器"
# 「01_管理确认」内仅允许本仓库生成器产出的文档格式
TOP_BUCKET_ALLOWED_SUFFIXES = frozenset({".md", ".json"})

# 与「第一批」同级：写在扫描根目录的父目录下，例如 资料根/第一批、资料根/检查结果
RESULT_DIR_NAME = "检查结果"

# 脚本写入各项目根目录的验收卡片长图（旧名「资料包…」含禁用词「资料」易被判不合格）
ACCEPTANCE_CARD_SCREENSHOT_PREFIX = "项目验收报告长图"
_LEGACY_ACCEPTANCE_CARD_SCREENSHOT_PREFIX = "资料包验收长图"
_RE_ACCEPTANCE_AUTO_LONG_PNG_STEM = re.compile(
    r"^(?:"
    + re.escape(ACCEPTANCE_CARD_SCREENSHOT_PREFIX)
    + "|"
    + re.escape(_LEGACY_ACCEPTANCE_CARD_SCREENSHOT_PREFIX)
    + r")_\d{8}_\d{6}$"
)

# 项目资料包根下第一层目录（管理 / 核心资料 / 安全）
# 根下必含；「03_安全与整改」不验收、不扫描、mock 不创建（见 SKIPPED_TOP_BUCKET_DIR）
PROJECT_TOP_LEVEL_DIRS: Tuple[str, ...] = (
    "01_管理确认",
    "02_核心资料",
)
SKIPPED_TOP_BUCKET_DIR = "03_安全与整改"
# 原「标准二级目录」01_～11_ 均落在「02_核心资料」之下；配置里 mock_std_dir 仍写叶子名（如 01_项目总览）。
CORE_MATERIALS_DIR = "02_核心资料"

# **核心资料区内**标准子目录全表 **01_～11_**（共 11 项）。第一阶段「缺目录」检查各规则 mock_std_dir（叶子名）落在 CORE_MATERIALS_DIR 下。
STANDARD_DIRS = [
    "01_项目总览",
    "02_售前方案",
    "03_业务需求",
    "04_产品设计",
    "05_实施上线",
    "06_验收交付",
    "07_项目复盘",
    "08_技术设计",
    "09_开发测试",
    "10_运维售后",
    "11_数据指标",
]

_allowed_std = set(STANDARD_DIRS)
for _r in PHASE1_RULES:
    _d = str(_r["mock_std_dir"])
    if _d not in _allowed_std:
        raise ValueError(
            f"acceptance_config 规则 {_r['id']}: mock_std_dir {_d!r} 不在 STANDARD_DIRS 中"
        )

# 第一阶段必验材料实际落位目录（去重、顺序与全表 STANDARD_DIRS 一致）
PHASE1_STANDARD_DIRS: Tuple[str, ...] = tuple(
    sorted(
        {str(r["mock_std_dir"]) for r in PHASE1_RULES},
        key=lambda d: STANDARD_DIRS.index(d),
    )
)
# 机检「须存在」的标准子目录（不含 optional 规则，如 r07/07_项目复盘允许旧包缺失）
PHASE1_REQUIRED_STANDARD_DIRS: Tuple[str, ...] = tuple(
    sorted(
        {str(r["mock_std_dir"]) for r in PHASE1_RULES if not r.get("optional")},
        key=lambda d: STANDARD_DIRS.index(d),
    )
)
PHASE1_LEAF_SET: Set[str] = set(PHASE1_STANDARD_DIRS)

FORBIDDEN_NAME_KEYWORDS = [
    "最终版",
    "最终终版",
    "最新版",
    "新建文档",
    "新建文件夹",
    "改过的",
    "客户资料",
    "资料",
    "文档",
    "备份",
]

RISK_KEYWORDS = [
    "接口文档",
    "数据字典",
    "部署手册",
    "安装配置",
    "报价",
    "合同",
    "账号",
    "密码",
    "服务器",
    "数据库",
    "样例数据",
    "巡检记录",
    "生产环境",
]

# 项目根目录文件夹命名规范（scan 字段仍为「项目文件夹命名是否合格」，报表列名带格式说明）
PROJECT_FOLDER_NAME_SPEC = "客户单位_项目名称_年份_项目状态"
REPORT_HEADER_PROJECT_FOLDER_OK = (
    "项目文件夹命名是否合格（须为「"
    + PROJECT_FOLDER_NAME_SPEC
    + "」；至少 4 段下划线分段，倒数第二段为四位年份）"
)
REPORT_HEADER_PROJECT_FOLDER_ISSUE = "项目文件夹命名问题说明"

# 六维闭环（业务目标 ↔ 脚本可自动核验部分；台账/签字等需人工）
CLOSURE_LEVEL_OK = "达标"
CLOSURE_LEVEL_WARN = "待加强"
CLOSURE_LEVEL_BAD = "需整改"

CLOSURE_DIMENSION_ORDER: Tuple[str, ...] = (
    "有资料",
    "资料归属清楚",
    "命名清楚",
    "目录清楚",
    "能追责",
    "能查找",
)


# =========================
# 2. 工具方法
# =========================


def normalize_name(name: str) -> str:
    """去掉空格，方便关键词匹配"""
    return name.replace(" ", "").replace("　", "").strip()


def list_all_files(project_dir: Path) -> List[Path]:
    """递归列出项目目录下所有文件（不进入「03_安全与整改」「确认单生成器」子树，与验收口径一致）。"""
    files: List[Path] = []
    project_dir = project_dir.resolve()
    for root, dirnames, filenames in os.walk(project_dir):
        if Path(root) == project_dir and SKIPPED_TOP_BUCKET_DIR in dirnames:
            dirnames.remove(SKIPPED_TOP_BUCKET_DIR)
        if CONFIRM_GENERATOR_DIRNAME in dirnames:
            dirnames.remove(CONFIRM_GENERATOR_DIRNAME)
        for filename in filenames:
            if filename.startswith("~$"):
                continue
            files.append(Path(root) / filename)
    return files


def check_project_folder_name(project_dir: Path) -> Tuple[bool, str]:
    """
    检查项目文件夹命名：
    客户单位_项目名称_年份_项目状态
    """
    name = project_dir.name
    parts = name.split("_")

    if len(parts) < 4:
        return (
            False,
            f"项目文件夹命名不规范，应为：{PROJECT_FOLDER_NAME_SPEC}（至少 4 段）",
        )

    year = parts[-2]
    if not re.fullmatch(r"\d{4}", year):
        return False, "项目文件夹年份不规范，应为 4 位年份，例如 2025"

    return True, ""


def check_standard_dirs(project_dir: Path) -> Tuple[List[str], List[str]]:
    """
    检查目录结构。

    - **缺失**：① 项目根下须具备 PROJECT_TOP_LEVEL_DIRS（01+02）；② 「02_核心资料」下须有各阶段所需标准叶子目录。
    - **额外**：项目根下除约定目录与可忽略的「03_安全与整改」外，不应多其他文件夹；「02_核心资料」下只允许 STANDARD_DIRS 中的叶子目录名。
    """
    root_names = [p.name for p in project_dir.iterdir() if p.is_dir()]

    missing_top = [d for d in PROJECT_TOP_LEVEL_DIRS if d not in root_names]

    core_path = project_dir / CORE_MATERIALS_DIR
    if core_path.is_dir():
        core_children = {p.name for p in core_path.iterdir() if p.is_dir()}
        missing_under_core = [
            d for d in PHASE1_REQUIRED_STANDARD_DIRS if d not in core_children
        ]
    else:
        missing_under_core = []

    missing_dirs: List[str] = list(missing_top)
    for leaf in missing_under_core:
        missing_dirs.append(f"{CORE_MATERIALS_DIR}/{leaf}")

    allowed_root = set(PROJECT_TOP_LEVEL_DIRS)
    extra_root = [
        d
        for d in root_names
        if d not in allowed_root and d != SKIPPED_TOP_BUCKET_DIR
    ]

    allowed_core_leaves = set(STANDARD_DIRS)
    extra_core: List[str] = []
    if core_path.is_dir():
        for name in sorted(p.name for p in core_path.iterdir() if p.is_dir()):
            if name not in allowed_core_leaves:
                extra_core.append(f"{CORE_MATERIALS_DIR}/{name}")

    extra_dirs = extra_root + extra_core

    return missing_dirs, extra_dirs


def _files_under_core_leaf(
    project_dir: Path, files: List[Path], leaf_name: str
) -> List[Path]:
    """列出已扫描到的、落在「02_核心资料/<leaf_name>/」下的文件（含子目录内文件）。"""
    out: List[Path] = []
    for f in files:
        try:
            rel = f.relative_to(project_dir).parts
        except ValueError:
            continue
        if (
            len(rel) >= 2
            and rel[0] == CORE_MATERIALS_DIR
            and rel[1] == leaf_name
        ):
            out.append(f)
    return out


def check_required_files(project_dir: Path, files: List[Path]) -> Dict[str, Dict[str, str]]:
    """
    检查第一阶段必验材料是否满足（规则见 PHASE1_RULES）。

    优先：对应「02_核心资料/<mock_std_dir>」下若已有任意文件，即视为「有」（按约定目录归档认定）。
    否则：仍按文件名是否含关键词命中（便于材料暂放他处时仍能识别）。
    """
    result: Dict[str, Dict[str, str]] = {}
    file_names = [normalize_name(f.name) for f in files]

    for rule in PHASE1_RULES:
        rid = str(rule["id"])
        std_dir = str(rule["mock_std_dir"])
        keywords: Sequence[str] = rule["keywords"]
        base_note = str(rule.get("note", ""))
        optional = bool(rule.get("optional"))

        in_leaf = _files_under_core_leaf(project_dir, files, std_dir)
        if in_leaf:
            note_extra = "约定目录内已有文件，按文件夹认定"
            note_out = f"{base_note}；{note_extra}" if base_note else note_extra
            result[rid] = {
                "label": str(rule["label"]),
                "status": "有",
                "matched_files": "; ".join(str(f) for f in in_leaf),
                "note": note_out,
            }
            continue

        matched: List[str] = []
        for f, normalized in zip(files, file_names):
            for keyword in keywords:
                if keyword in normalized:
                    matched.append(str(f))
                    break

        if optional:
            st = "有" if matched else "缺（可选）"
            if matched:
                note_out = base_note
            else:
                note_out = f"{base_note}（旧项目可无）" if base_note else "旧项目可无"
        else:
            st = "有" if matched else "缺"
            note_out = base_note

        result[rid] = {
            "label": str(rule["label"]),
            "status": st,
            "matched_files": "; ".join(matched),
            "note": note_out,
        }

    return result


def file_under_bucket(file_path: Path, project_dir: Path, bucket: str) -> bool:
    """判定文件是否落在项目根下指定第一层分区目录内。"""
    try:
        rel = file_path.relative_to(project_dir)
    except ValueError:
        return False
    parts = rel.parts
    return len(parts) >= 1 and parts[0] == bucket


def check_top_bucket_files(
    project_dir: Path,
    files: List[Path],
) -> Dict[str, Dict[str, str]]:
    """
    顶层分区（如 01_管理确认，见 TOP_BUCKET_RULES）资料检查：
    仅在对应分区内、且扩展名为 .md / .json 的文件上扫描文件名关键词；「03_安全与整改」已整体跳过、不验。
    """
    result: Dict[str, Dict[str, str]] = {}
    if not TOP_BUCKET_RULES:
        return result

    for rule in TOP_BUCKET_RULES:
        rid = str(rule["id"])
        bucket = str(rule["bucket_dir"])
        keywords: Sequence[str] = tuple(rule["keywords"])
        matched: List[str] = []
        for f in files:
            if not file_under_bucket(f, project_dir, bucket):
                continue
            if f.suffix.lower() not in TOP_BUCKET_ALLOWED_SUFFIXES:
                continue
            nn = normalize_name(f.name)
            for keyword in keywords:
                if keyword in nn:
                    matched.append(str(f))
                    break
        result[rid] = {
            "label": str(rule["label"]),
            "status": "有" if matched else "缺",
            "matched_files": "; ".join(matched),
            "note": str(rule.get("note", "")),
            "分区目录": bucket,
        }
    return result


def _rel_parts_project(file_path: Path, project_dir: Optional[Path]) -> Tuple[str, ...]:
    if project_dir is None:
        return tuple()
    try:
        return file_path.relative_to(project_dir).parts
    except ValueError:
        return tuple()


def _under_core_phase1_leaf(rel: Tuple[str, ...]) -> bool:
    return (
        len(rel) >= 3
        and rel[0] == CORE_MATERIALS_DIR
        and rel[1] in PHASE1_LEAF_SET
    )


def _under_top_bucket(rel: Tuple[str, ...]) -> bool:
    return len(rel) >= 2 and rel[0] == "01_管理确认"


def _legacy_seven_segment_check(stem: str) -> Tuple[str, List[str]]:
    """完整七段命名（历史规则）。"""
    issues: List[str] = []
    parts = stem.split("_")
    if len(parts) < 7:
        issues.append(
            "命名字段不足，应为：客户单位_项目名称_资料类别_资料名称_版本号_日期_密级"
        )
        return "不合格", issues

    category = parts[2]
    version = parts[-3]
    date = parts[-2]
    secret = parts[-1]

    if category not in VALID_CATEGORIES:
        issues.append(f"资料类别不规范：{category}")

    if not VERSION_SEGMENT_RE.fullmatch(version):
        issues.append(f"版本号不规范：{version}")

    if not DATE_SEGMENT_RE.fullmatch(date):
        issues.append(f"日期格式不规范：{date}，应为 YYYYMMDD")
    else:
        try:
            datetime.strptime(date, "%Y%m%d")
        except ValueError:
            issues.append(f"日期不是有效日期：{date}")

    if secret not in VALID_SECRET_LEVELS:
        issues.append(f"密级不规范：{secret}")

    risk_hit = [kw for kw in RISK_KEYWORDS if kw in stem]
    if risk_hit:
        if secret in {"公开", "内部"} and "脱敏版" not in stem:
            issues.append(
                f"疑似敏感资料未标记脱敏或密级较低，命中关键词：{','.join(risk_hit)}"
            )

    if not issues:
        return "合格", []
    if len(parts) == 7:
        return "警告", issues
    return "不合格", issues


def _simplified_three_segment_check(stem: str) -> Tuple[str, List[str]]:
    """资料名称_日期_密级（三段式）。"""
    parsed = try_parse_simplified_stem(stem)
    if parsed is None:
        return (
            "不合格",
            [
                "须为：资料名称_日期_密级（日期8位YYYYMMDD，密级：公开/内部/敏感/核心）",
            ],
        )
    _doc, _dt, secret = parsed
    issues: List[str] = []
    risk_hit = [kw for kw in RISK_KEYWORDS if kw in stem]
    if risk_hit:
        if secret in {"公开", "内部"} and "脱敏版" not in stem:
            issues.append(
                f"疑似敏感资料未标记脱敏或密级较低，命中关键词：{','.join(risk_hit)}"
            )
    if not issues:
        return "合格", []
    return "警告", issues


def _confirmation_bucket_two_segment_check(stem: str) -> Tuple[str, List[str]]:
    """01_管理确认：资料名称_YYYYMMDD。"""
    if try_parse_confirmation_bucket_stem(stem) is None:
        return "不合格", ["须为：资料名称_YYYYMMDD（末段为8位日期）"]
    issues: List[str] = []
    risk_hit = [kw for kw in RISK_KEYWORDS if kw in stem]
    if risk_hit and "脱敏版" not in stem:
        issues.append(
            f"文件名命中敏感/风险词：{','.join(risk_hit)}，建议在文件名中标注「脱敏版」或调整密级归档方式"
        )
    if not issues:
        return "合格", []
    return "警告", issues


def _explain_confirmation_bucket_rejection(stem: str) -> str:
    """两段式解析失败时的补充说明。"""
    parts = stem.split("_")
    if len(parts) < 2:
        return "（须至少两段：资料名称_YYYYMMDD）"
    date = parts[-1]
    doc = "_".join(parts[:-1])
    msgs: List[str] = []
    if not doc.strip():
        msgs.append("资料名称不能为空")
    if try_parse_simplified_stem(stem) is not None:
        msgs.append("「01_管理确认」已改为两段式，请去掉末段密级")
    elif date.isdigit() and len(date) == 6:
        msgs.append("末段日期须为8位YYYYMMDD（勿仅用6位年月）")
    elif not DATE_SEGMENT_RE.fullmatch(date):
        msgs.append("末段须为8位YYYYMMDD")
    else:
        try:
            datetime.strptime(date, "%Y%m%d")
        except ValueError:
            msgs.append("末段日期不是有效日历日")
    if msgs:
        return "（" + "；".join(msgs) + "）"
    return ""


def _explain_simplified_stem_rejection(stem: str) -> str:
    """try_parse_simplified_stem 失败时补充可读原因（如日期误写成 YYYYMM）。"""
    parts = stem.split("_")
    if len(parts) < 3:
        return ""
    doc = "_".join(parts[:-2])
    date = parts[-2]
    secret = parts[-1]
    msgs: List[str] = []
    if not doc.strip():
        msgs.append("资料名称不能为空")
    if secret not in VALID_SECRET_LEVELS:
        msgs.append(f"末段密级须为：{'、'.join(sorted(VALID_SECRET_LEVELS))}")
    if date.isdigit() and len(date) == 6:
        msgs.append("日期须为8位YYYYMMDD（勿仅用6位年月YYYYMM，可改为如20250401）")
    elif not DATE_SEGMENT_RE.fullmatch(date):
        msgs.append("日期须为8位YYYYMMDD")
    else:
        try:
            datetime.strptime(date, "%Y%m%d")
        except ValueError:
            msgs.append("日期不是有效日历日")
    if msgs:
        return "（" + "；".join(msgs) + "）"
    return ""


def _is_project_root_acceptance_long_png(
    file_path: Path, project_dir: Optional[Path]
) -> bool:
    """验收脚本生成的汇总卡片长图，位于项目根目录，不参与资料命名规则考核。"""
    if project_dir is None or file_path.suffix.lower() != ".png":
        return False
    try:
        if file_path.parent.resolve() != project_dir.resolve():
            return False
    except OSError:
        return False
    return _RE_ACCEPTANCE_AUTO_LONG_PNG_STEM.fullmatch(file_path.stem) is not None


def _remove_prior_acceptance_long_pngs(proj_dir: Path) -> None:
    """删除项目根目录下历史验收长图，仅保留本轮即将写入的新文件。"""
    try:
        for p in proj_dir.iterdir():
            if (
                p.is_file()
                and p.suffix.lower() == ".png"
                and _RE_ACCEPTANCE_AUTO_LONG_PNG_STEM.fullmatch(p.stem)
            ):
                p.unlink()
    except OSError:
        pass


def check_file_name(
    file_path: Path,
    project_dir: Optional[Path] = None,
) -> Tuple[str, List[str]]:
    """
    文件命名检查。

    • 「01_管理确认」：**仅允许 .md / .json**；命名 **资料名称_YYYYMMDD** 两段式（末段日期，无密级）。
    • 「02_核心资料」下**七类**叶子：**资料名称_日期_密级** 三段式（只校格式，**不**验资料名称是否含必交关键词）。
    • 「03_安全与整改」不扫描。
    • 「确认单生成器」工具目录不扫描（见 list_all_files）。
    • 其余路径：完整七段（见 naming_convention）。
    """
    base: List[str] = []
    stem = file_path.stem
    suffix = file_path.suffix.lower()

    rel = _rel_parts_project(file_path, project_dir)

    if _is_project_root_acceptance_long_png(file_path, project_dir):
        return "合格", []

    if project_dir is not None and rel and _under_top_bucket(rel):
        if suffix not in TOP_BUCKET_ALLOWED_SUFFIXES:
            suf_disp = suffix if suffix else "无后缀"
            return "不合格", [
                f"「01_管理确认」内仅允许 .md、.json 文件，当前为「{suf_disp}」"
            ]
    elif suffix not in VALID_EXTENSIONS:
        base.append(f"文件后缀不在允许范围：{suffix}")

    for word in FORBIDDEN_NAME_KEYWORDS:
        if stem == word or stem.startswith(word):
            base.append(f"文件名包含或等于禁用命名：{word}")

    if project_dir is not None and rel and (
        _under_core_phase1_leaf(rel) or _under_top_bucket(rel)
    ):
        if _under_top_bucket(rel):
            if try_parse_simplified_stem(stem) is not None:
                return "不合格", base + [
                    "「01_管理确认」内须使用「资料名称_YYYYMMDD」两段式（无密级段）；请去掉末段密级或改名"
                ]
            if try_parse_confirmation_bucket_stem(stem) is not None:
                st, naming_issues = _confirmation_bucket_two_segment_check(stem)
            else:
                hint = "「01_管理确认」内须使用「资料名称_YYYYMMDD」（末段为8位日期）"
                return "不合格", base + [
                    hint + _explain_confirmation_bucket_rejection(stem)
                ]
        elif try_parse_simplified_stem(stem) is not None:
            st, naming_issues = _simplified_three_segment_check(stem)
        else:
            hint = "「02_核心资料」下七类目录内须使用「资料名称_日期_密级」"
            return "不合格", base + [hint + _explain_simplified_stem_rejection(stem)]
    else:
        st, naming_issues = _legacy_seven_segment_check(stem)

    all_issues = base + naming_issues
    if not all_issues:
        return "合格", []

    if st == "不合格":
        return "不合格", all_issues

    return "警告", all_issues


def calculate_level(
    naming_pass_rate: float,
    required_complete_rate: float,
    high_risk_count: int,
    missing_required_count: int,
) -> str:
    """根据 `acceptance_config.ACCEPTANCE_THRESHOLDS` 给出初步验收建议。"""
    a = ACCEPTANCE_THRESHOLDS["level_a"]
    if (
        naming_pass_rate >= a["naming_pass_min"]
        and required_complete_rate >= a["required_complete_min"]
        and missing_required_count <= a.get("max_missing_required", 0)
        and (not a.get("require_zero_high_risk") or high_risk_count == 0)
    ):
        return "A-建议入库"

    b = ACCEPTANCE_THRESHOLDS["level_b"]
    if (
        naming_pass_rate >= b["naming_pass_min"]
        and required_complete_rate >= b["required_complete_min"]
        and missing_required_count <= b.get("max_missing_required", 999)
    ):
        return "B-小修后入库"

    c = ACCEPTANCE_THRESHOLDS["level_c"]
    if (
        naming_pass_rate >= c["naming_pass_min"]
        and required_complete_rate >= c["required_complete_min"]
    ):
        return "C-退回整改"

    return "D-暂缓入库"


def evaluate_closure_six(r: Dict[str, Any]) -> Dict[str, Dict[str, str]]:
    """
    将「有资料、归属、命名、目录、追责、可查找」六维与现有扫描指标对齐，
    每项给出 达标 / 待加强 / 需整改 + 一句话说明（不含台账/签字等纯人工项）。
    """
    missing_required = int(r.get("必交资料缺失数") or 0)
    folder_ok = r.get("项目文件夹命名是否合格") == "是"
    folder_issue = str(r.get("项目文件夹命名问题") or "")
    missing_dirs: List[str] = list(r.get("第一阶段标准目录缺失") or [])
    extra_dirs: List[str] = list(r.get("额外目录") or [])
    np = float(r.get("命名合格率") or 0.0)
    nf = int(r.get("命名不合格数") or 0)
    nw = int(r.get("命名警告数") or 0)
    hr = int(r.get("风险提示文件数") or 0)

    out: Dict[str, Dict[str, str]] = {}

    # 1. 有资料
    if missing_required == 0:
        out["有资料"] = {
            "status": CLOSURE_LEVEL_OK,
            "note": "核心资料必交（约定目录或关键词）与分区资料均已满足（仅核文件名与目录）",
        }
    else:
        out["有资料"] = {
            "status": CLOSURE_LEVEL_BAD,
            "note": f"有 {missing_required} 项（核心必交或分区资料）未命中，需补件或调整文件名",
        }

    # 2. 资料归属清楚
    if folder_ok and np >= 85.0:
        out["资料归属清楚"] = {
            "status": CLOSURE_LEVEL_OK,
            "note": "项目根目录含客户/项目/年份/状态；多数文件可按规范文件名（含三段式）识别归属",
        }
    elif folder_ok:
        out["资料归属清楚"] = {
            "status": CLOSURE_LEVEL_WARN,
            "note": "根目录规范；部分文件命名未达规范，归属识别度一般",
        }
    else:
        out["资料归属清楚"] = {
            "status": CLOSURE_LEVEL_BAD,
            "note": folder_issue or "项目根目录无法稳定识别客户与项目",
        }

    # 3. 命名清楚
    if nf == 0 and np >= 90.0:
        out["命名清楚"] = {
            "status": CLOSURE_LEVEL_OK,
            "note": "命名合格率高，无「不合格」档文件",
        }
    elif nf == 0 and np >= 70.0:
        out["命名清楚"] = {
            "status": CLOSURE_LEVEL_WARN,
            "note": f"无不合格命名；存在 {nw} 条警告，建议收敛版本/日期/密级等字段",
        }
    elif nf == 0:
        out["命名清楚"] = {
            "status": CLOSURE_LEVEL_WARN,
            "note": "无不合格命名，但整体合格率偏低，建议统一规范",
        }
    else:
        out["命名清楚"] = {
            "status": CLOSURE_LEVEL_BAD,
            "note": f"有 {nf} 个文件命名不合格，影响识别与批量管理",
        }

    # 4. 目录清楚
    if not missing_dirs and not extra_dirs:
        out["目录清楚"] = {
            "status": CLOSURE_LEVEL_OK,
            "note": f"根下 01+02 齐，「{CORE_MATERIALS_DIR}」下标准子目录齐，无多余文件夹",
        }
    elif missing_dirs:
        out["目录清楚"] = {
            "status": CLOSURE_LEVEL_BAD,
            "note": "缺标准子目录：" + "；".join(missing_dirs),
        }
    else:
        out["目录清楚"] = {
            "status": CLOSURE_LEVEL_WARN,
            "note": "存在非约定目录或多余文件夹："
            + "；".join(extra_dirs)
            + "（建议归位到「"
            + CORE_MATERIALS_DIR
            + "」或登记说明）",
        }

    # 5. 能追责（依命名要素 + 风险提示；签字流程不在脚本范围）
    if folder_ok and hr == 0 and np >= 75.0:
        out["能追责"] = {
            "status": CLOSURE_LEVEL_OK,
            "note": "根目录与文件名可追溯客户/项目/版本/日期/密级；未触发敏感风险提示",
        }
    elif hr > 0:
        out["能追责"] = {
            "status": CLOSURE_LEVEL_WARN,
            "note": f"有 {hr} 个文件命中敏感/风险关键词，需人工确认脱密与密级标注",
        }
    elif not folder_ok:
        out["能追责"] = {
            "status": CLOSURE_LEVEL_BAD,
            "note": folder_issue or "项目根目录无法稳定对应责任边界",
        }
    else:
        out["能追责"] = {
            "status": CLOSURE_LEVEL_WARN,
            "note": "命名尚可；建议结合台账与交付清单强化责任链路",
        }

    # 6. 能查找（目录 + 命名 + 必交的合成结论）
    if (
        missing_required == 0
        and not missing_dirs
        and nf == 0
        and np >= 80.0
        and folder_ok
    ):
        out["能查找"] = {
            "status": CLOSURE_LEVEL_OK,
            "note": "分层目录 + 规范命名 + 必交齐全，便于按路径与关键词检索",
        }
    elif missing_required > 0 or missing_dirs:
        out["能查找"] = {
            "status": CLOSURE_LEVEL_BAD,
            "note": "缺材料或缺标准目录时，检索入口不完整",
        }
    elif not folder_ok or np < 60.0:
        out["能查找"] = {
            "status": CLOSURE_LEVEL_BAD,
            "note": "根目录或命名混乱时，检索与定位成本高",
        }
    else:
        out["能查找"] = {
            "status": CLOSURE_LEVEL_WARN,
            "note": "基本可找件；建议继续统一命名，并配合单位检索台账/标签体系",
        }

    return out


def safe_relative_path(path: Path, base: Path) -> str:
    try:
        return str(path.relative_to(base))
    except ValueError:
        return str(path)


# =========================
# 3. 核心扫描逻辑
# =========================


def scan_project(project_dir: Path, root_dir: Path) -> Dict[str, Any]:
    project_name = project_dir.name

    folder_ok, folder_issue = check_project_folder_name(project_dir)
    missing_dirs, extra_dirs = check_standard_dirs(project_dir)
    files = list_all_files(project_dir)

    required_result = check_required_files(project_dir, files)
    top_bucket_result = check_top_bucket_files(project_dir, files)

    file_check_rows: List[Dict[str, str]] = []
    naming_ok_count = 0
    naming_warning_count = 0
    naming_fail_count = 0
    high_risk_count = 0

    for file in files:
        status, issues = check_file_name(file, project_dir)

        if status == "合格":
            naming_ok_count += 1
        elif status == "警告":
            naming_warning_count += 1
        else:
            naming_fail_count += 1

        if any("疑似敏感资料" in issue for issue in issues):
            high_risk_count += 1

        file_check_rows.append(
            {
                "项目名称": project_name,
                "文件路径": safe_relative_path(file, project_dir),
                "文件名": file.name,
                "检查结果": status,
                "问题说明": "；".join(issues),
            }
        )

    total_files = len(files)
    naming_pass_rate = (
        round((naming_ok_count / total_files) * 100, 2) if total_files else 0.0
    )

    # 核心资料区（PHASE1 规则，不含 optional）
    required_total = sum(1 for r in PHASE1_RULES if not r.get("optional"))
    if required_total == 0:
        core_have = 0
        core_missing = 0
        core_rate = 100.0
    else:
        core_have = sum(
            1
            for r in PHASE1_RULES
            if not r.get("optional")
            and required_result[str(r["id"])]["status"] == "有"
        )
        core_missing = required_total - core_have
        core_rate = round((core_have / required_total) * 100, 2)

    # 顶层分区（配置项，当前不含 03）
    tb_total = len(top_bucket_result)
    if tb_total == 0:
        tb_have = 0
        tb_missing = 0
        tb_rate = 100.0
    else:
        tb_have = sum(
            1 for item in top_bucket_result.values() if item["status"] == "有"
        )
        tb_missing = tb_total - tb_have
        tb_rate = round((tb_have / tb_total) * 100, 2)

    merge_total = required_total + tb_total
    if merge_total == 0:
        merge_have = 0
        missing_required_count = 0
        required_complete_rate = 100.0
    else:
        merge_have = core_have + tb_have
        missing_required_count = merge_total - merge_have
        required_complete_rate = round((merge_have / merge_total) * 100, 2)

    level = calculate_level(
        naming_pass_rate=naming_pass_rate,
        required_complete_rate=required_complete_rate,
        high_risk_count=high_risk_count,
        missing_required_count=missing_required_count,
    )

    row = {
        "项目名称": project_name,
        "项目路径": safe_relative_path(project_dir, root_dir),
        "项目文件夹命名是否合格": "是" if folder_ok else "否",
        "项目文件夹命名问题": folder_issue,
        "第一阶段标准目录缺失": missing_dirs,
        "额外目录": extra_dirs,
        "文件总数": total_files,
        "命名合格数": naming_ok_count,
        "命名警告数": naming_warning_count,
        "命名不合格数": naming_fail_count,
        "命名合格率": naming_pass_rate,
        "核心资料必交完整率": core_rate,
        "核心资料必交缺失数": core_missing,
        "分区资料完整率": tb_rate,
        "分区资料缺失数": tb_missing,
        "必交资料完整率": required_complete_rate,
        "必交资料缺失数": missing_required_count,
        "风险提示文件数": high_risk_count,
        "验收建议": level,
        "必交资料检查": required_result,
        "分区资料检查": top_bucket_result,
        "文件命名检查": file_check_rows,
    }
    row["六维闭环"] = evaluate_closure_six(row)
    return row


def scan_root(root_dir: Path) -> List[Dict[str, Any]]:
    projects: List[Path] = []

    for item in root_dir.iterdir():
        if item.is_dir():
            # 跳过批次根下「00_」类汇总等非项目包目录（勿跳过以 01_/02_ 开头的合法项目文件夹名）
            if item.name.startswith("00_"):
                continue
            # 若误将「检查结果」放在批次目录内，勿当作项目包
            if item.name == RESULT_DIR_NAME:
                continue
            projects.append(item)

    results: List[Dict[str, Any]] = []
    for project in projects:
        results.append(scan_project(project, root_dir))

    return results


# =========================
# 4. Excel 报告生成
# =========================

_THIN = Side(style="thin", color="B4B4B4")
_GRID_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill(
    fill_type="solid", start_color="1F4E79", end_color="1F4E79"
)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_DATA_FONT = Font(size=11)
_ZEBRA_FILL = PatternFill(fill_type="solid", start_color="F5F5F5", end_color="F5F5F5")


def _visual_text_units(s: str) -> float:
    """估算单元格显示宽度（拉丁字符约 1，中日韩等约 2）。"""
    u = 0.0
    for ch in s:
        u += 2.05 if ord(ch) > 127 else 1.0
    return u


def auto_column_widths(ws: Any, max_width_by_col: Dict[int, float]) -> None:
    """按内容估算列宽，避免过长列撑爆表格。"""
    default_cap = 52.0
    for col in ws.columns:
        idx = col[0].column
        letter = get_column_letter(idx)
        max_u = 12.0
        for cell in col:
            max_u = max(
                max_u,
                _visual_text_units(str(cell.value if cell.value is not None else "")),
            )
        cap = max_width_by_col.get(idx, default_cap)
        ws.column_dimensions[letter].width = min(max(max_u * 0.52 + 2.8, 11), cap)


def format_report_sheet(
    ws: Any,
    *,
    wrap_columns: Tuple[int, ...] = (),
    percent_columns: Tuple[int, ...] = (),
    status_ok_col: Optional[int] = None,
    level_col: Optional[int] = None,
    check_result_col: Optional[int] = None,
    closure_status_cols: Tuple[int, ...] = (),
) -> None:
    """冻结首行、筛选、边框、斑马纹、长文本换行、关键列底色。"""
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.row_dimensions[1].height = 38

    wrap_set = set(wrap_columns)

    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _GRID_BORDER

    for r in range(2, max_row + 1):
        zebra = r % 2 == 0
        needs_tall = False
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _DATA_FONT
            cell.border = _GRID_BORDER
            wrap = c in wrap_set
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=wrap,
            )
            if wrap:
                needs_tall = True
            fill_for_cell = None
            if zebra:
                fill_for_cell = _ZEBRA_FILL
            if status_ok_col == c:
                v = str(cell.value or "")
                if v == "有":
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="E8F5E9",
                        end_color="E8F5E9",
                    )
                elif v == "缺":
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="FFEBEE",
                        end_color="FFEBEE",
                    )
                elif v == "缺（可选）":
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="F1F5F9",
                        end_color="F1F5F9",
                    )
            if level_col == c:
                v = str(cell.value or "")
                if v.startswith("A-"):
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="C6EFCE",
                        end_color="C6EFCE",
                    )
                elif v.startswith("B-"):
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="FFEB9C",
                        end_color="FFEB9C",
                    )
                elif v.startswith("C-"):
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="FCE4D6",
                        end_color="FCE4D6",
                    )
                elif v.startswith("D-"):
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="F4B084",
                        end_color="F4B084",
                    )
            if check_result_col == c:
                v = str(cell.value or "")
                if v == "合格":
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="E8F5E9",
                        end_color="E8F5E9",
                    )
                elif v == "警告":
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="FFF9E6",
                        end_color="FFF9E6",
                    )
                elif v == "不合格":
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="FFEBEE",
                        end_color="FFEBEE",
                    )
            if c in closure_status_cols:
                v = str(cell.value or "")
                if v == CLOSURE_LEVEL_OK:
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="E8F5E9",
                        end_color="E8F5E9",
                    )
                elif v == CLOSURE_LEVEL_WARN:
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="FFF9E6",
                        end_color="FFF9E6",
                    )
                elif v == CLOSURE_LEVEL_BAD:
                    fill_for_cell = PatternFill(
                        fill_type="solid",
                        start_color="FFEBEE",
                        end_color="FFEBEE",
                    )
            if fill_for_cell is not None:
                cell.fill = fill_for_cell

            if c in percent_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '0.00"%"'

        if needs_tall:
            h = ws.row_dimensions[r].height
            if h is None or (isinstance(h, (int, float)) and h < 22):
                ws.row_dimensions[r].height = 22


def write_report(results: List[Dict[str, Any]], output_path: Path) -> None:
    wb = Workbook()

    ws_guide = wb.active
    ws_guide.title = "阅读说明"
    ws_guide.column_dimensions["A"].width = 96
    ws_guide["A1"] = "资料包自动验收报告 · 阅读说明"
    ws_guide["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws_guide["A3"] = (
        "「总览」每个项目一行：命名合格率 / 必交资料完整率为百分比；"
        "「验收建议」列按 A～D 档着色便于扫读。"
    )
    ws_guide["A4"] = (
        "「必交资料检查」每条规则一行：绿色=已命中（约定目录内有文件，或文件名含关键词），红色=缺失。"
    )
    ws_guide["A5"] = (
        "「文件命名检查」逐文件列出结果：合格 / 警告 / 不合格列已浅色区分。"
    )
    ws_guide["A6"] = (
        "「整改清单」汇总待处理项。各表均已冻结首行、开启筛选，长文本列自动换行。"
    )
    ws_guide["A7"] = (
        "「总览」中「项目文件夹命名是否合格」列：须为「"
        + PROJECT_FOLDER_NAME_SPEC
        + "」，至少 4 段下划线、倒数第二段为四位年份。"
    )
    ws_guide["A8"] = (
        "【六维闭环】与「有资料、归属清、命名清、目录清、能追责、能查找」对齐；"
        "「六维闭环」表中每项为「"
        + CLOSURE_LEVEL_OK
        + " / "
        + CLOSURE_LEVEL_WARN
        + " / "
        + CLOSURE_LEVEL_BAD
        + "」。"
        "①有资料：必交项是否在约定目录内有文件，或文件名命中关键词（不读正文）。②资料归属：项目根目录 + 规范文件名能否识别客户/项目。"
        "③命名清楚：合格/警告/不合格统计。④目录清楚：根下 01+02 是否齐、"
        f"「{CORE_MATERIALS_DIR}」下标准叶子目录是否齐、是否有额外文件夹。"
        "⑤能追责：版本/日期/密级与风险提示（签字审批走台账，脚本不验）。⑥能查找：①～⑤合成；台账标签属人工协同。"
        "详见「六维闭环」工作表。"
    )
    ws_guide["A9"] = (
        "【资料包根目录结构】须直接包含："
        + "、".join(PROJECT_TOP_LEVEL_DIRS)
        + "。原标准目录名（01_项目总览～11_数据指标）均建在「"
        + CORE_MATERIALS_DIR
        + "」下；缺失/额外目录列中的路径均相对项目根。"
        "「分区资料检查」工作表对应 acceptance_config.TOP_BUCKET_RULES。"
        "「01_管理确认」：仅 .md / .json；命名「资料名称_YYYYMMDD」两段式；「02_核心资料」七类：「资料名称_日期_密级」三段式。必交是否齐见「必交资料检查」（约定目录内有文件即可认定，或文件名关键词）。"
        "详见 tobacco_kb.naming_convention。"
    )
    for rr in (3, 4, 5, 6, 7, 8, 9):
        ws_guide.cell(rr, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws_guide.row_dimensions[1].height = 30

    ws = wb.create_sheet("总览", 1)

    headers = [
        "项目名称",
        "项目路径",
        "文件总数",
        "命名合格数",
        "命名警告数",
        "命名不合格数",
        "命名合格率",
        "核心资料必交完整率",
        "核心资料必交缺失数",
        "分区资料完整率（01）",
        "分区资料缺失数（01）",
        "合计必交完整率（核心+分区）",
        "合计必交缺失数（核心+分区）",
        "风险提示文件数",
        REPORT_HEADER_PROJECT_FOLDER_OK,
        REPORT_HEADER_PROJECT_FOLDER_ISSUE,
        "第一阶段标准目录缺失",
        "额外目录",
        "验收建议",
    ]
    ws.append(headers)

    for r in results:
        ws.append(
            [
                r["项目名称"],
                r["项目路径"],
                r["文件总数"],
                r["命名合格数"],
                r["命名警告数"],
                r["命名不合格数"],
                r["命名合格率"],
                r["核心资料必交完整率"],
                r["核心资料必交缺失数"],
                r["分区资料完整率"],
                r["分区资料缺失数"],
                r["必交资料完整率"],
                r["必交资料缺失数"],
                r["风险提示文件数"],
                r["项目文件夹命名是否合格"],
                r["项目文件夹命名问题"],
                "；".join(r["第一阶段标准目录缺失"]),
                "；".join(r["额外目录"]),
                r["验收建议"],
            ]
        )

    # 总览：长文列；百分比列为命名合格率、三类完整率
    auto_column_widths(
        ws,
        {
            1: 22,
            2: 36,
            3: 12,
            4: 12,
            5: 12,
            6: 12,
            7: 14,
            8: 18,
            9: 18,
            10: 18,
            11: 18,
            12: 22,
            13: 22,
            14: 16,
            15: 52,
            16: 44,
            17: 40,
            18: 28,
            19: 22,
        },
    )
    format_report_sheet(
        ws,
        wrap_columns=(2, 15, 16, 17, 18, 19),
        percent_columns=(7, 8, 10, 12),
        level_col=19,
    )

    ws_closure = wb.create_sheet("六维闭环", 2)
    closure_headers = ["项目名称", *CLOSURE_DIMENSION_ORDER, "综合说明"]
    ws_closure.append(closure_headers)
    for r in results:
        closure = r.get("六维闭环") or {}
        line_parts: List[str] = []
        row_cells: List[Any] = [r["项目名称"]]
        for dim in CLOSURE_DIMENSION_ORDER:
            item = closure.get(dim, {})
            st = str(item.get("status", ""))
            row_cells.append(st)
            line_parts.append(
                f"{dim}【{st}】{item.get('note', '')}"
            )
        row_cells.append("\n".join(line_parts))
        ws_closure.append(row_cells)
    auto_column_widths(
        ws_closure,
        {
            1: 22,
            2: 11,
            3: 13,
            4: 11,
            5: 11,
            6: 11,
            7: 11,
            8: 52,
        },
    )
    format_report_sheet(
        ws_closure,
        wrap_columns=(8,),
        closure_status_cols=tuple(range(2, 8)),
    )

    ws2 = wb.create_sheet("必交资料检查")
    ws2.append(
        [
            "项目名称",
            "必交资料项",
            "是否存在",
            "匹配到的文件",
            "材料说明",
        ]
    )

    for r in results:
        for rule in PHASE1_RULES:
            rid = str(rule["id"])
            item = r["必交资料检查"][rid]
            ws2.append(
                [
                    r["项目名称"],
                    item["label"],
                    item["status"],
                    item["matched_files"],
                    item.get("note", ""),
                ]
            )

    auto_column_widths(
        ws2,
        {1: 22, 2: 38, 3: 12, 4: 58, 5: 36},
    )
    format_report_sheet(
        ws2,
        wrap_columns=(2, 4, 5),
        status_ok_col=3,
    )

    ws_tb = wb.create_sheet("分区资料检查")
    ws_tb.append(
        [
            "项目名称",
            "分区目录",
            "资料项",
            "是否存在",
            "匹配到的文件",
            "说明",
        ]
    )
    for r in results:
        tb_chk = r.get("分区资料检查") or {}
        for rule in TOP_BUCKET_RULES:
            rid = str(rule["id"])
            item = tb_chk.get(rid)
            if item is None:
                continue
            ws_tb.append(
                [
                    r["项目名称"],
                    item.get("分区目录", ""),
                    item["label"],
                    item["status"],
                    item["matched_files"],
                    item.get("note", ""),
                ]
            )
    auto_column_widths(
        ws_tb,
        {1: 22, 2: 16, 3: 36, 4: 12, 5: 52, 6: 36},
    )
    format_report_sheet(
        ws_tb,
        wrap_columns=(3, 5, 6),
        status_ok_col=4,
    )

    ws3 = wb.create_sheet("文件命名检查")
    ws3.append(
        [
            "项目名称",
            "文件路径",
            "文件名",
            "检查结果",
            "问题说明",
        ]
    )

    for r in results:
        for row in r["文件命名检查"]:
            ws3.append(
                [
                    row["项目名称"],
                    row["文件路径"],
                    row["文件名"],
                    row["检查结果"],
                    row["问题说明"],
                ]
            )

    auto_column_widths(ws3, {1: 22, 2: 42, 3: 28, 4: 14, 5: 56})
    format_report_sheet(
        ws3,
        wrap_columns=(2, 5),
        check_result_col=4,
    )

    ws4 = wb.create_sheet("整改清单")
    ws4.append(
        [
            "项目名称",
            "问题类型",
            "问题描述",
            "整改建议",
        ]
    )

    for r in results:
        if r["项目文件夹命名是否合格"] == "否":
            ws4.append(
                [
                    r["项目名称"],
                    "项目文件夹命名",
                    r["项目文件夹命名问题"],
                    f"按「{PROJECT_FOLDER_NAME_SPEC}」重新命名",
                ]
            )

        if r["第一阶段标准目录缺失"]:
            ws4.append(
                [
                    r["项目名称"],
                    "第一阶段标准目录缺失",
                    "缺少：" + "；".join(r["第一阶段标准目录缺失"]),
                    "补齐第一阶段所需标准子目录（见 tobacco_kb.acceptance_config），若无资料可保留空目录",
                ]
            )

        for rule in PHASE1_RULES:
            rid = str(rule["id"])
            item = r["必交资料检查"][rid]
            if item["status"] == "缺":
                ws4.append(
                    [
                        r["项目名称"],
                        "必交资料缺失",
                        f"缺少：{item['label']}",
                        "在规则对应标准目录内放入资料，或使文件名含配置关键词（见 mock_std_dir 与 keywords）；如确实没有，备注说明",
                    ]
                )

        tb_chk = r.get("分区资料检查") or {}
        for rule in TOP_BUCKET_RULES:
            rid = str(rule["id"])
            item = tb_chk.get(rid)
            if item is None:
                continue
            if item["status"] == "缺":
                ws4.append(
                    [
                        r["项目名称"],
                        "分区资料缺失（01）",
                        f"「{item.get('分区目录', '')}」缺少：{item['label']}",
                        "在对应分区目录内放置文件，文件名须含配置关键词之一（见 acceptance_config.TOP_BUCKET_RULES）",
                    ]
                )

        for row in r["文件命名检查"]:
            if row["检查结果"] in {"警告", "不合格"}:
                ws4.append(
                    [
                        r["项目名称"],
                        "文件命名问题",
                        f"{row['文件路径']}：{row['问题说明']}",
                        "按命名约定重命名：「01_管理确认」仅 .md / .json，且为「资料名称_YYYYMMDD」两段式；「02_核心资料」七类为「资料名称_日期_密级」三段式；其余路径见 tobacco_kb.naming_convention（完整七段）",
                    ]
                )

    auto_column_widths(ws4, {1: 22, 2: 22, 3: 54, 4: 44})
    format_report_sheet(ws4, wrap_columns=(3, 4))

    wb.save(output_path)


def write_report_html(results: List[Dict[str, Any]], output_path: Path) -> None:
    """将验收结果写成单文件 HTML（与 Excel 同源数据，便于浏览器查看）。"""

    def e(x: Any) -> str:
        return html.escape(str(x if x is not None else ""), quote=True)

    title = "资料包自动验收报告"
    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 与 calculate_level 返回值一致，用于 HTML 着色
    level_class = {
        "A-建议入库": "lvl-a",
        "B-小修后入库": "lvl-b",
        "C-退回整改": "lvl-c",
        "D-暂缓入库": "lvl-d",
    }

    def bar_pct(val: Any) -> str:
        try:
            v = float(val)
        except (TypeError, ValueError):
            return "0"
        return str(max(0.0, min(100.0, v)))

    parts: List[str] = []
    parts.append("<!DOCTYPE html>\n")
    parts.append('<html lang="zh-CN">\n<head>\n')
    parts.append('<meta charset="utf-8">\n')
    parts.append(
        '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
    )
    parts.append(f"<title>{e(title)}</title>\n")
    parts.append("<style>\n")
    parts.append(
        ":root{--bg:#f5f7fb;--card:#fff;--bd:#e2e8f0;--txt:#0f172a;"
        "--muted:#64748b;--ok:#16a34a;--warn:#ca8a04;--bad:#dc2626;"
        "--bar-bg:#e2e8f0;}\n"
        "*{box-sizing:border-box;}\n"
        "body{font-family:system-ui,-apple-system,'Segoe UI',Roboto,'Microsoft YaHei',sans-serif;"
        "margin:0;background:var(--bg);color:var(--txt);line-height:1.55;}\n"
        ".wrap{max-width:1200px;margin:0 auto;padding:20px 16px 48px;}\n"
        "header{background:linear-gradient(135deg,#1e3a5f,#0f172a);color:#fff;"
        "padding:20px 24px;border-radius:12px;margin-bottom:20px;}\n"
        "header h1{margin:0 0 8px;font-size:1.35rem;font-weight:700;}\n"
        ".meta{color:#cbd5e1;font-size:0.9rem;}\n"
        "h2{font-size:1.1rem;margin:24px 0 12px;padding-bottom:6px;"
        "border-bottom:2px solid var(--bd);}\n"
        "table{border-collapse:collapse;width:100%;background:var(--card);"
        "border-radius:10px;overflow:hidden;box-shadow:0 1px 3px rgba(15,23,42,.06);}\n"
        "th,td{border:1px solid var(--bd);padding:8px 10px;font-size:0.875rem;"
        "vertical-align:top;}\n"
        "thead th{background:#f1f5f9;text-align:left;}\n"
        "tbody tr:nth-child(even){background:#fafbfc;}\n"
        ".num{text-align:right;white-space:nowrap;}\n"
        ".name{font-weight:600;}\n"
        ".scroll-x{overflow-x:auto;-webkit-overflow-scrolling:touch;}\n"
        ".lvl-a{color:var(--ok);font-weight:600;}\n"
        ".lvl-b{color:var(--warn);font-weight:600;}\n"
        ".lvl-c,.lvl-d{color:var(--bad);font-weight:600;}\n"
        ".lvl-unknown{color:var(--muted);}\n"
        ".bar{height:8px;background:var(--bar-bg);border-radius:999px;overflow:hidden;"
        "margin-top:4px;}\n"
        ".bar>span{display:block;height:100%;background:#3b82f6;border-radius:999px;}\n"
        "nav.toc{background:var(--card);padding:12px 14px;border-radius:10px;"
        "border:1px solid var(--bd);margin-bottom:20px;font-size:0.9rem;}\n"
        "nav.toc a{color:#2563eb;margin-right:12px;text-decoration:none;}\n"
        "nav.toc a:hover{text-decoration:underline;}\n"
        ".project-card{background:var(--card);border:1px solid var(--bd);"
        "border-radius:12px;padding:16px 18px;margin-bottom:18px;"
        "box-shadow:0 1px 3px rgba(15,23,42,.06);}\n"
        ".project-card h3{margin:0 0 10px;font-size:1.05rem;}\n"
        ".path{color:var(--muted);font-size:0.82rem;word-break:break-all;"
        "margin-bottom:12px;}\n"
        ".metrics{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));"
        "gap:10px;margin-bottom:14px;}\n"
        ".metric{background:#f8fafc;border:1px solid var(--bd);border-radius:8px;"
        "padding:8px 10px;}\n"
        ".metric .k{font-size:0.75rem;color:var(--muted);}\n"
        ".metric .hint{font-size:0.72rem;color:var(--muted);font-weight:400;"
        "margin-top:4px;line-height:1.4;}\n"
        ".metric .v{font-size:0.95rem;font-weight:600;margin-top:6px;}\n"
        ".req-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));"
        "gap:8px;margin:10px 0;}\n"
        ".req{display:flex;flex-direction:column;border:1px solid var(--bd);"
        "border-radius:8px;padding:8px 10px;background:#fafbfc;font-size:0.82rem;}\n"
        ".req .rid{color:var(--muted);font-size:0.72rem;margin-bottom:4px;}\n"
        ".req.ok{border-left:4px solid var(--ok);}\n"
        ".req.miss{border-left:4px solid var(--bad);}\n"
        ".req.opt-miss{border-left:4px solid #94a3b8;background:#f8fafc;}\n"
        ".tag{display:inline-block;padding:2px 8px;border-radius:999px;font-size:0.75rem;"
        "font-weight:600;margin-top:6px;}\n"
        ".tag-ok{background:#dcfce7;color:#166534;}\n"
        ".tag-miss{background:#fee2e2;color:#991b1b;}\n"
        ".tag-opt{background:#e2e8f0;color:#475569;}\n"
        ".chk-ok{color:var(--ok);font-weight:600;}\n"
        ".chk-warn{color:var(--warn);font-weight:600;}\n"
        ".chk-bad{color:var(--bad);font-weight:600;}\n"
        "details{margin:10px 0;border:1px solid var(--bd);border-radius:8px;"
        "padding:8px 12px;background:#fff;}\n"
        "summary{cursor:pointer;font-weight:600;font-size:0.9rem;}\n"
        "ul.rect{margin:8px 0 0 18px;padding:0;}\n"
        "ul.rect li{margin:6px 0;font-size:0.85rem;}\n"
        "details.rectify-summary{margin:12px 0;border:2px solid #dc2626;border-radius:8px;"
        "padding:10px 12px;background:#fff1f2;box-shadow:0 1px 4px rgba(220,38,38,.18);}\n"
        "details.rectify-summary>summary{cursor:pointer;color:#b91c1c;font-weight:800;"
        "font-size:0.98rem;letter-spacing:0.03em;}\n"
        "details.rectify-summary ul.rect{margin-top:10px;}\n"
        "details.rectify-summary ul.rect li{color:#7f1d1d;}\n"
        ".closure-banner{background:var(--card);border:1px solid var(--bd);border-radius:10px;"
        "padding:14px 16px;margin-bottom:20px;font-size:0.85rem;line-height:1.65;}\n"
        ".closure-banner strong{color:#1e3a5f;}\n"
        ".closure-h{font-size:0.95rem;margin:0 0 10px;color:#1e3a5f;}\n"
        ".closure-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(230px,1fr));"
        "gap:10px;margin-bottom:8px;}\n"
        ".cl-cell{border:1px solid var(--bd);border-radius:8px;padding:10px 12px;font-size:0.82rem;"
        "background:#fafbfc;}\n"
        ".cl-cell .dn{font-weight:600;margin-bottom:6px;color:#475569;font-size:0.76rem;}\n"
        ".cl-cell .st{font-weight:700;margin-bottom:6px;font-size:0.88rem;}\n"
        ".cl-cell .no{color:var(--muted);font-size:0.76rem;line-height:1.45;}\n"
        ".st-ok .st{color:var(--ok);}\n"
        ".st-warn .st{color:var(--warn);}\n"
        ".st-bad .st{color:var(--bad);}\n"
        "footer{margin-top:28px;color:var(--muted);font-size:0.8rem;text-align:center;}\n"
        "</style>\n</head>\n<body>\n<div class=\"wrap\">\n"
    )

    def _closure_cell_class(st: str) -> str:
        if st == CLOSURE_LEVEL_OK:
            return "cl-cell st-ok"
        if st == CLOSURE_LEVEL_WARN:
            return "cl-cell st-warn"
        return "cl-cell st-bad"

    parts.append("<header><h1>")
    parts.append(e(title))
    parts.append("</h1>")
    parts.append(
        f'<div class="meta">生成时间：{e(generated)} · '
        f"共 {len(results)} 个项目资料包</div></header>\n"
    )

    parts.append('<section class="closure-banner">\n')
    parts.append("<strong>六维闭环（业务目标）</strong>：有资料 · 资料归属清楚 · 命名清楚 · ")
    parts.append("目录清楚 · 能追责 · 能查找。")
    parts.append(
        "脚本依据文件名关键词、标准目录、命名规则、风险提示等给出「"
        + CLOSURE_LEVEL_OK
        + " / "
        + CLOSURE_LEVEL_WARN
        + " / "
        + CLOSURE_LEVEL_BAD
        + "」；台账签字等须人工协同。详见 Excel「六维闭环」表及各项目下卡片。\n"
    )
    parts.append("</section>\n")

    parts.append('<h2>总览</h2>\n<div class="scroll-x"><table><thead><tr>')
    parts.append(
        "<th>#</th><th>项目名称</th><th>项目路径</th><th>命名合格率</th>"
        "<th>合计必交完整率（核心+分区）</th>"
        "<th>验收建议（A～D 档）</th>"
        "</tr></thead><tbody>\n"
    )
    for idx, r in enumerate(results, start=1):
        suggestion = r.get("验收建议", "")
        lc = level_class.get(str(suggestion), "lvl-unknown")
        nm_pct = bar_pct(r.get("命名合格率"))
        req_pct = bar_pct(r.get("必交资料完整率"))
        parts.append("<tr>")
        parts.append(f'<td class="num">{idx}</td>')
        parts.append(f'<td class="name">{e(r.get("项目名称",""))}</td>')
        parts.append(f'<td>{e(r.get("项目路径",""))}</td>')
        parts.append("<td>")
        parts.append(f'<div>{e(r.get("命名合格率",""))}</div>')
        parts.append(
            f'<div class="bar"><span style="width:{e(nm_pct)}%"></span></div>'
        )
        parts.append("</td><td>")
        parts.append(f'<div>{e(r.get("必交资料完整率",""))}</div>')
        parts.append(
            f'<div class="bar"><span style="width:{e(req_pct)}%"></span></div>'
        )
        parts.append("</td>")
        parts.append(f'<td class="{lc}">{e(suggestion)}</td>')
        parts.append("</tr>\n")
    parts.append("</tbody></table></div>\n")

    parts.append('<nav class="toc"><strong>快速跳转：</strong>')
    for idx, r in enumerate(results, start=1):
        nm = r.get("项目名称", f"项目{idx}")
        parts.append(f'<a href="#p{idx}">{e(nm)}</a>')
    parts.append("</nav>\n")

    parts.append("<h2>各项目详情</h2>\n")
    for idx, r in enumerate(results, start=1):
        suggestion = r.get("验收建议", "")
        lc = level_class.get(str(suggestion), "lvl-unknown")
        parts.append(f'<section class="project-card" id="p{idx}">\n')
        parts.append(f"<h3>{e(r.get('项目名称',''))}</h3>\n")
        parts.append(f'<div class="path">{e(r.get("项目路径",""))}</div>\n')

        closure = r.get("六维闭环") or {}
        parts.append('<p class="closure-h">六维闭环检视</p>\n')
        parts.append('<div class="closure-grid">\n')
        for dim in CLOSURE_DIMENSION_ORDER:
            item = closure.get(dim, {})
            st = str(item.get("status", ""))
            cc = _closure_cell_class(st)
            parts.append(f'<div class="{e(cc)}">\n')
            parts.append(f'<div class="dn">{e(dim)}</div>\n')
            parts.append(f'<div class="st">{e(st)}</div>\n')
            parts.append(f'<div class="no">{e(item.get("note", ""))}</div>\n')
            parts.append("</div>\n")
        parts.append("</div>\n")

        parts.append('<div class="metrics">\n')
        parts.append(
            f'<div class="metric"><div class="k">命名合格率</div>'
            f'<div class="v">{e(r.get("命名合格率",""))}</div></div>\n'
        )
        parts.append(
            f'<div class="metric"><div class="k">核心资料必交完整率</div>'
            f'<div class="v">{e(r.get("核心资料必交完整率",""))}</div></div>\n'
        )
        parts.append(
            f'<div class="metric"><div class="k">分区资料完整率（01）</div>'
            f'<div class="v">{e(r.get("分区资料完整率",""))}</div></div>\n'
        )
        parts.append(
            f'<div class="metric"><div class="k">合计必交完整率</div>'
            f'<div class="v">{e(r.get("必交资料完整率",""))}</div></div>\n'
        )
        parts.append('<div class="metric">\n')
        parts.append('<div class="k">验收建议（A～D 档）</div>\n')
        parts.append(
            '<div class="hint">综合命名合格率、核心+分区必交完整率、合计缺失条数、'
            "风险提示文件数等，按 acceptance_config 阈值自动分档</div>\n"
        )
        parts.append(
            f'<div class="v {lc}">{e(suggestion)}</div>\n'
        )
        parts.append("</div>\n")
        parts.append('<div class="metric">\n')
        parts.append('<div class="k">项目文件夹命名是否合格</div>\n')
        parts.append(
            f'<div class="hint">须为「{e(PROJECT_FOLDER_NAME_SPEC)}」；'
            f"至少 4 段下划线、倒数第二段四位年份</div>\n"
        )
        parts.append(
            f'<div class="v">{e(r.get("项目文件夹命名是否合格",""))}</div>\n'
        )
        parts.append("</div>\n")
        parts.append("</div>\n")

        req_chk = r.get("必交资料检查") or {}
        parts.append(
            '<p style="margin:12px 0 6px;font-weight:600;color:#1e3a5f;">'
                "核心资料必交（约定标准目录内有文件，或文件名关键词）</p>\n"
        )
        parts.append('<div class="req-grid">\n')
        for rule in PHASE1_RULES:
            rid = str(rule["id"])
            item = req_chk.get(rid, {})
            status = item.get("status", "")
            miss = status == "缺"
            opt_miss = status == "缺（可选）"
            if miss:
                css = "req miss"
                tag_cls = "tag-miss"
                tag_txt = "缺"
            elif opt_miss:
                css = "req opt-miss"
                tag_cls = "tag-opt"
                tag_txt = "可选缺"
            else:
                css = "req ok"
                tag_cls = "tag-ok"
                tag_txt = "有"
            parts.append(f'<div class="{css}">\n')
            parts.append(f'<div class="rid">{e(rid)} · {e(rule.get("label",""))}</div>\n')
            parts.append(f'<div>{e(item.get("label",""))}</div>\n')
            parts.append(f'<span class="tag {tag_cls}">{e(tag_txt)}</span>\n')
            if item.get("matched_files"):
                parts.append(
                    f'<div style="margin-top:8px;color:var(--muted);font-size:0.78rem;">'
                    f'命中：{e(item.get("matched_files",""))}</div>\n'
                )
            if item.get("note"):
                parts.append(
                    f'<div style="margin-top:6px;color:var(--muted);font-size:0.78rem;">'
                    f'备注：{e(item.get("note",""))}</div>\n'
                )
            parts.append("</div>\n")
        parts.append("</div>\n")

        tb_chk = r.get("分区资料检查") or {}
        if TOP_BUCKET_RULES:
            parts.append(
                '<p style="margin:16px 0 6px;font-weight:600;color:#1e3a5f;">'
                "分区资料（「01_管理确认」内仅 .md/.json + 文件名关键词；03 不验）</p>\n"
            )
            parts.append('<div class="req-grid">\n')
            for rule in TOP_BUCKET_RULES:
                rid = str(rule["id"])
                item = tb_chk.get(rid, {})
                status = item.get("status", "")
                miss = status == "缺"
                css = "req miss" if miss else "req ok"
                parts.append(f'<div class="{css}">\n')
                parts.append(
                    f'<div class="rid">{e(rid)} · {e(item.get("分区目录",""))}</div>\n'
                )
                parts.append(f'<div>{e(item.get("label",""))}</div>\n')
                tag_cls = "tag-miss" if miss else "tag-ok"
                tag_txt = "缺" if miss else "有"
                parts.append(f'<span class="tag {tag_cls}">{e(tag_txt)}</span>\n')
                if item.get("matched_files"):
                    parts.append(
                        f'<div style="margin-top:8px;color:var(--muted);font-size:0.78rem;">'
                        f'命中：{e(item.get("matched_files",""))}</div>\n'
                    )
                if item.get("note"):
                    parts.append(
                        f'<div style="margin-top:6px;color:var(--muted);font-size:0.78rem;">'
                        f'备注：{e(item.get("note",""))}</div>\n'
                    )
                parts.append("</div>\n")
            parts.append("</div>\n")

        parts.append("<details open>\n<summary>目录检查（缺失 / 额外）</summary>\n")
        miss_dirs = r.get("第一阶段标准目录缺失") or []
        extra_dirs = r.get("额外目录") or []
        parts.append("<p><strong>第一阶段标准目录缺失：</strong>")
        parts.append(e("；".join(miss_dirs)) if miss_dirs else "无")
        parts.append("</p>\n")
        parts.append("<p><strong>根目录下额外目录：</strong>")
        parts.append(e("；".join(extra_dirs)) if extra_dirs else "无")
        parts.append("</p>\n</details>\n")

        if r.get("项目文件夹命名是否合格") == "否":
            parts.append(
                "<details>\n<summary>项目文件夹命名问题说明（规范同上）</summary>\n"
            )
            parts.append(f'<p>{e(r.get("项目文件夹命名问题",""))}</p>\n</details>\n')

        parts.append(
            "<details open class=\"rectify-summary\">\n"
            "<summary>整改清单（摘要）</summary>\n"
            "<ul class=\"rect\">\n"
        )
        any_rect = False
        if r.get("项目文件夹命名是否合格") == "否":
            any_rect = True
            parts.append(
                "<li><strong>项目文件夹命名：</strong>"
                f"{e(r.get('项目文件夹命名问题',''))} — "
                f"按「{PROJECT_FOLDER_NAME_SPEC}」重新命名</li>\n"
            )
        for d in miss_dirs:
            any_rect = True
            parts.append(
                "<li><strong>第一阶段标准目录缺失：</strong>"
                f"缺少：{e(d)} — 补齐标准子目录（无可保留空目录）</li>\n"
            )
        for rule in PHASE1_RULES:
            rid = str(rule["id"])
            item = req_chk.get(rid, {})
            if item.get("status") == "缺":
                any_rect = True
                parts.append(
                    "<li><strong>必交资料缺失：</strong>"
                    f"缺少：{e(item.get('label',''))} — "
                    "在对应标准目录内放资料，或文件名含规则关键词；如确无，备注说明</li>\n"
                )
        tb_chk = r.get("分区资料检查") or {}
        for rule in TOP_BUCKET_RULES:
            rid = str(rule["id"])
            item = tb_chk.get(rid)
            if item is None:
                continue
            if item.get("status") == "缺":
                any_rect = True
                parts.append(
                    "<li><strong>分区资料缺失（01）：</strong>"
                    f"「{e(item.get('分区目录',''))}」{e(item.get('label',''))} — "
                    "在对应目录下放文件，文件名含关键词之一（见 TOP_BUCKET_RULES）</li>\n"
                )
        for row in r.get("文件命名检查") or []:
            if row.get("检查结果") in {"警告", "不合格"}:
                any_rect = True
                parts.append(
                    "<li><strong>文件命名问题：</strong>"
                    f"{e(row.get('文件路径',''))}：{e(row.get('问题说明',''))} — "
                    "按命名约定重命名（01 两段式、02 七类三段式，见 naming_convention）</li>\n"
                )
        if not any_rect:
            parts.append("<li>当前无自动生成的整改项。</li>\n")
        parts.append("</ul>\n</details>\n")

        naming_rows = r.get("文件命名检查") or []
        parts.append("<details open>\n<summary>文件命名检查（")
        parts.append(str(len(naming_rows)))
        parts.append(" 条）</summary>\n")
        if not naming_rows:
            parts.append("<p>无文件记录。</p>\n")
        else:
            parts.append('<div class="scroll-x"><table><thead><tr>')
            parts.append(
                "<th>文件路径</th><th>文件名</th><th>检查结果</th><th>问题说明</th>"
                "</tr></thead><tbody>\n"
            )
            for row in naming_rows:
                chk = row.get("检查结果", "")
                cc = "chk-ok"
                if chk == "警告":
                    cc = "chk-warn"
                elif chk == "不合格":
                    cc = "chk-bad"
                parts.append("<tr>")
                parts.append(f'<td>{e(row.get("文件路径",""))}</td>')
                parts.append(f'<td>{e(row.get("文件名",""))}</td>')
                parts.append(f'<td class="{cc}">{e(chk)}</td>')
                parts.append(f'<td>{e(row.get("问题说明",""))}</td>')
                parts.append("</tr>\n")
            parts.append("</tbody></table></div>\n")
        parts.append("</details>\n")

        parts.append("</section>\n")

    parts.append(
        f'<footer>本页由验收脚本自动生成 · {e(generated)} · '
        "数据与同目录 Excel 一致（含「六维闭环」工作表）</footer>\n"
    )
    parts.append("</div>\n</body>\n</html>\n")

    output_path.write_text("".join(parts), encoding="utf-8")


# =========================
# 5. 命令行入口
# =========================


def _configure_stdio_utf8() -> None:
    if sys.platform != "win32":
        return
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8")
        except Exception:
            pass


def export_per_project_check_long_pngs(
    root_dir: Path,
    results: List[Dict[str, Any]],
    batch_html_path: Path,
    stamp: str,
) -> None:
    """根据汇总 HTML 中各项目卡片（#p1…#pN）截取整卡长图，写入对应项目资料包目录。"""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print(
            "提示：未安装 playwright，已跳过各项目验收长图 PNG。"
            "安装：pip install playwright，然后执行：playwright install chromium"
        )
        return

    targets: List[Tuple[int, Path]] = []
    for idx, r in enumerate(results, start=1):
        rel = str(r.get("项目路径") or "").strip()
        if not rel:
            continue
        proj_dir = root_dir / rel
        if not proj_dir.is_dir():
            continue
        _remove_prior_acceptance_long_pngs(proj_dir)
        targets.append(
            (idx, proj_dir / f"{ACCEPTANCE_CARD_SCREENSHOT_PREFIX}_{stamp}.png")
        )

    if not targets:
        return

    html_uri = batch_html_path.resolve().as_uri()
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page(viewport={"width": 1280, "height": 900})
        page.goto(html_uri, wait_until="load", timeout=120_000)
        for idx, png_path in targets:
            loc = page.locator(f"#p{idx}")
            loc.scroll_into_view_if_needed()
            loc.screenshot(path=str(png_path))
        browser.close()

    print(
        f"已在各项目资料包目录生成验收长图 PNG（共 {len(targets)} 个，与汇总网页中项目卡片一致）。"
    )


def main() -> None:
    _configure_stdio_utf8()
    # 默认：当前工作目录下的「第一批」；也可传入自定义根目录覆盖
    if len(sys.argv) >= 2:
        root_dir = Path(sys.argv[1]).resolve()
    else:
        root_dir = (Path.cwd() / "第一批").resolve()

    if not root_dir.exists():
        print(f"目录不存在：{root_dir}")
        if len(sys.argv) < 2:
            print("请在含有「第一批」文件夹的目录下执行本脚本，或显式传入路径，例如：")
            print(
                r'python check_tobacco_kb_required_files.py "D:\资料上报\第一批"'
            )
        sys.exit(1)

    if not root_dir.is_dir():
        print(f"传入路径不是目录：{root_dir}")
        sys.exit(1)

    print(f"开始扫描目录：{root_dir}")

    results = scan_root(root_dir)

    if not results:
        print("未发现项目资料包，请确认根目录下是否存在项目文件夹。")
        sys.exit(0)

    report_dir = root_dir.parent / RESULT_DIR_NAME
    report_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = report_dir / f"资料包自动验收报告_{stamp}.xlsx"
    output_html = report_dir / f"资料包自动验收报告_{stamp}.html"

    write_report(results, output_path)
    write_report_html(results, output_html)
    export_per_project_check_long_pngs(root_dir, results, output_html, stamp)

    print("扫描完成。")
    print(f"验收报告（Excel）已生成：{output_path}")
    print(f"验收报告（网页）已生成：{output_html}")


if __name__ == "__main__":
    main()
